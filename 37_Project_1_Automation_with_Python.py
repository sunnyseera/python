print()

# An application to process excel spreadsheets

# Import a package as an alias to make our code shorter
import openpyxl as xl 

# Import bar chart so we can visualse the data
from openpyxl.chart import BarChart, Reference
from openpyxl.xml.constants import MAX_COLUMN

# Open the Excel Sheet 
wb = xl.load_workbook('transactions.xlsx')

# Open a specific sheet in the workbook
sheet = wb['Sheet1']

# Select a cell to start from within the sheet
cell = sheet['a1']

# This will print the value of CELL A1
print(cell.value)

#################################################################

# Calcualte the amount of rows and columns in the sheet
print(sheet.max_row)
print()

# Loop through rows 2 to the max to ingonore the header 
for row in range(2, sheet.max_row + 1) : 

    # This will print the values in the rows
    cell = sheet.cell(row, 3)
    print(cell.value)

    # Now we will ammend the prices to remove 10% of the value 
    # and add it to a new column
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# Select all of the values in the 4th Column
Values = Reference(  sheet, 
            min_row=2, 
            max_row=sheet.max_row, 
            min_col=4, 
            max_col=4
)

# Add the Chart 
# Create the Object 
chart = BarChart()
# Add the data 
chart.add_data(Values)
#Specify the sheet and the location of the chart 
sheet.add_chart(chart, 'E2')

# Save the changes to the workbook
wb.save('transactions_corrected.xlsx')

#################################################################